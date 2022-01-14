from django.http.response import HttpResponse
from django.shortcuts import get_object_or_404, render
from .models import FileUpLoad, TestModel
from django_pandas.io import pd
import io
from django.http import JsonResponse
import matplotlib.pyplot as plt
#from django.http import HttpResponse
import base64
import matplotlib
matplotlib.use('Agg')
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
import os
from django.core import serializers
import openpyxl
import ast

#画面更新を行うコード。直接エクセルに射影する
def calltest(req,pk):
    if req.method == 'GET':
        file_value=get_object_or_404(FileUpLoad,id=pk)
        message_a=ast.literal_eval(req.GET.get("txt"))
        subtest=str(req.GET.get("subtest")) #検索用 シート名
        test=str(req.GET.get("test"))       #検索用 テスト名
        process=str(req.GET.get("process")) #検索用 工程名

        wb = openpyxl.load_workbook(file_value.upload_dir.path, keep_vba=True) #keep_vbaが無いとxlsm拡張子にもかかわらず中身はxlsxで保存され、破損ファイル扱いになる
        ws=wb.get_sheet_by_name(subtest)

        #print(subtest)
        if len(message_a)!=0:
            if test=="" and process=="":
                coordinates=list(message_a.keys()) #変更箇所の座標を取得(CommonSettingを(0,0)として定義)
                for coor in coordinates:
                    if coor.split(",")[0]!=0: #x座標が0ならデータフレームにしか存在しない列なので変更しない。
                        updatemsg = message_a[coor]
                        x,y=(int(item) for item in coor.split(","))
                        #xはエクセル上でそのまま、yはスキップした分を足す。
                        x += 7
                        ws.cell(x,y).value = updatemsg.replace("\n","")
                context={"data":"SUCCESS"}
                wb.save(file_value.upload_dir.path)
                return JsonResponse(context,safe = False)
            else:
                #この辺りに、テスト名と工程名とCommonSetting名でエクセル抽出した結果を挿入して、インデックスをmotherfileと合わせるコードを追加する
                xlsrow = [] #xlsrow[x-1]がtableの行インデックスxに対応
                for i in range(8,ws.max_row):
                    if test.lower() in str(ws.cell(i,2).value).lower() and process.lower() in str(ws.cell(i,4).value).lower():
                        xlsrow.append(i)
                coordinates=list(message_a.keys()) #変更箇所の座標を取得(CommonSettingを(0,0)として定義)
                for coor in coordinates:
                    if coor.split(",")[0]!=0: #x座標が0ならデータフレームにしか存在しない列なので変更しない。
                        updatemsg = message_a[coor]
                        x,y=(int(item) for item in coor.split(","))
                        #xはエクセル上でそのまま、yはスキップした分を足す。
                        #x += 7
                        ws.cell(xlsrow[x-1],y).value = updatemsg.replace("\n","")
                context={"data":"SUCCESS"}
                wb.save(file_value.upload_dir.path)
                return JsonResponse(context,safe = False)
        else:
            #更新箇所が存在しない場合のケア。トランザクションの競合も対策する。
            content={"data":"NO UPDATE"}
            return JsonResponse(content,safe=False)


# Create your views here.
#viewにて普通のプログラミングっぽいのが可能
def test2(req):
    #all()はpythonの組み込みメソッド。オブジェクト(class内で定義した変数たち)を持ってきて、all()で全て格納する。
    file_obj=FileUpLoad.objects.all()
    #辞書式のリスト。
    context={
        'file_obj':file_obj,
    }
    return render(req,'test2.html',context)
    #reqが発生した際に、htmlにcontext(modelsで定義した)を受け渡すのがrender.
def testfunction(req):
    if req.method == 'GET':
        testname=req.GET.get("txt1")
        testnumber=req.GET.get("txt2")
        if testnumber=="":
            testnumber=-100
        test=TestModel(testname=testname,testnumber=testnumber)
        test.save()
        context={
            'data':TestModel.objects.all(),
            'data2':TestModel.objects.values_list("testname",flat=True).distinct()
        }
        print(context['data2'])
        return render(req,'testmodel.html',context)

#はじめにdetail.htmlに遷移したとき、ファイル情報を読むコード
def detail(req,pk):
    file_value=get_object_or_404(FileUpLoad,id=pk)
    if os.path.splitext(file_value.upload_dir.path)[1]==".csv":
        try:
            df=pd.read_csv(file_value.upload_dir.path,index_col=0,skip_blank_lines=True,skiprows=8)
        except UnicodeDecodeError:
            df=pd.read_csv(file_value.upload_dir.path,index_col=0,skip_blank_lines=True,encoding='cp932',skiprows=8)
    else:
        df=pd.read_excel(file_value.upload_dir.path,skiprows=6,sheet_name=0)
        input_file = pd.ExcelFile(file_value.upload_dir.path)
        sheets = input_file.sheet_names
        lists=['No','TestName', 'A-C', 'Process', 'Min', 'Typ', 'Max', 'Unit', 'Value',
               'Pin', 'Type', 'Parameter', 'Value.1', 'Unit.1', 'Pin.1', 'Measurement',
               'Parameter.1', 'Value.2', 'Unit.2', 'Type.1', 'Item', 'Ope', 'Add',
               'Data', 'MeasPoint', 'Comment', 'Comment.1']
        df=df[lists]
        #df=df.dropna(how="all")
        df=df.fillna("-")
        pair=""
        df['No']=df['No'].replace('-','UNKNOWNVALUE')
        #-------------------------------------------------
        wb = openpyxl.load_workbook(file_value.upload_dir.path) #rowspan用のセル結合抜き出し
        ws = wb[sheets[0]]
        cellcolor = ws.cell(8,2).fill.bgColor.value
        cellcontcolor = ws.cell(1,1).fill.bgColor.value
        start=0
        for i in range(8, ws.max_row):
            if str(ws.cell(i,2).value).lower() == "common setting":
                if start == 0:
                    pair={}
                    start=1
                elif start == 1:
                    pair.update({commonname:{"skip":count, "content":count_cont}})
                commonname=ws.cell(i+1,2).value
                count=1
                count_cont=0
            elif str(ws.cell(i,2).value).lower() != "common setting" and ws.cell(i,2).fill.bgColor.value == cellcolor:
                count+=1
            elif ws.cell(i,2).fill.bgColor.value == cellcontcolor and ws.cell(i+1,4).fill.bgColor.value != cellcontcolor:
                count_cont+=1
            elif ws.cell(i,2).fill.bgColor.value == cellcontcolor and ws.cell(i+1,4).fill.bgColor.value == cellcontcolor:
                count_cont+=1
                pair.update({commonname:{"skip":count, "content":count_cont}})
                break
        df["CommonSetting"]="---"
        lists_2=list(pair.keys())      #common_setting名の配列
        k=0
        for i in range(len(lists_2)):
            #df.iat[k,len(df.columns)-1]=lists[i]
            skipnum = pair[lists_2[i]]["skip"]+pair[lists_2[i]]["content"]
            df.iloc[k:k+skipnum,len(df.columns)-1]=lists_2[i]
            k+=skipnum
        #-------------------------------------------------
        df=df.reindex(columns=[
            'CommonSetting','No','TestName', 'A-C', 'Process', 'Min', 'Typ', 'Max', 'Unit', 'Value',
            'Pin', 'Type', 'Parameter', 'Value.1', 'Unit.1', 'Pin.1', 'Measurement',
            'Parameter.1', 'Value.2', 'Unit.2', 'Type.1', 'Item', 'Ope', 'Add',
            'Data', 'MeasPoint', 'Comment', 'Comment.1'
        ])
    context={
            'file_value':file_value, #csv/excelデータ
            'df':df,                 #表に記載するデータ
            'sheets':sheets,         #サブテスト名
            'cs':pair,                #Common Settingのセル結合数の辞書式リスト
            'default':sheets[0]
    }
    return render(req,'detail.html',context)



#検索結果と一致するデータフレームのみ抽出してjsonresponseで送信する
def detail_search(req,pk):
    if req.method == 'GET':
        search = str(req.GET.get('input_data'))
        process = str(req.GET.get('input_data2'))
        subtest = str(req.GET.get('input_data3'))
        searchcommon = str(req.GET.get('input_data4'))
        file_value=get_object_or_404(FileUpLoad,id=pk)
        if os.path.splitext(file_value.upload_dir.path)[1]==".csv":
            try:
                df=pd.read_csv(file_value.upload_dir.path,index_col=0,skip_blank_lines=True,skiprows=8)
            except UnicodeDecodeError:
                df=pd.read_csv(file_value.upload_dir.path,index_col=0,skip_blank_lines=True,encoding='cp932',skiprows=8)
        else:
            #df=pd.read_excel(file_value.upload_dir.path,skiprows=6,sheet_name=1)
            #df=pd.read_excel(file_value.upload_dir.path,skiprows=6)
            input_file = pd.ExcelFile(file_value.upload_dir.path)
            sheets = input_file.sheet_names
            sheet_read=""
            df=pd.read_excel(file_value.upload_dir.path,skiprows=6,sheet_name=1)
            for sheet in sheets:
                if (subtest.lower() in sheet.lower()) and subtest!="サブテスト選択" and subtest!="ALL":
                    sheet_read=sheet
                    index=sheets.index(sheet)
                    df=pd.read_excel(file_value.upload_dir.path,skiprows=6,sheet_name=sheet)
                    break
                elif subtest=="ALL":
                    sheetlist=[]
                    sheet_read="ALL"
                    for i in sheets:
                        df_alpha=pd.read_excel(file_value.upload_dir.path,skiprows=6, sheet_name=i)
                        sheetlist.append(df_alpha)
                    df=pd.concat(sheetlist)
                    break
            lists=['No','TestName', 'A-C', 'Process', 'Min', 'Typ', 'Max', 'Unit', 'Value',
                'Pin', 'Type', 'Parameter', 'Value.1', 'Unit.1', 'Pin.1', 'Measurement',
                'Parameter.1', 'Value.2', 'Unit.2', 'Type.1', 'Item', 'Ope', 'Add',
                'Data', 'MeasPoint', 'Comment', 'Comment.1']
            df=df[lists]
            df=df.dropna(how="all")
            df=df.fillna("-")
            #df['No']=df['No'].replace('-','UNKNOWNVALUE')

            #-------------------------------------------------
            wb = openpyxl.load_workbook(file_value.upload_dir.path) #rowspan用のセル結合抜き出し
            ws = wb[sheets[index]]
            cellcolor = ws.cell(8,2).fill.bgColor.value
            cellcontcolor = ws.cell(1,1).fill.bgColor.value
            start=0
            for i in range(8, ws.max_row):
                if str(ws.cell(i,2).value).lower() == "common setting":
                    if start == 0:
                        pair={}
                        start=1
                    elif start == 1:
                        pair.update({commonname:{"skip":count, "content":count_cont}})
                    commonname=ws.cell(i+1,2).value
                    count=1
                    count_cont=0
                elif str(ws.cell(i,2).value).lower() != "common setting" and ws.cell(i,2).fill.bgColor.value == cellcolor:
                    count+=1
                elif ws.cell(i,2).fill.bgColor.value == cellcontcolor and ws.cell(i+1,4).fill.bgColor.value != cellcontcolor:
                    count_cont+=1
                elif ws.cell(i,2).fill.bgColor.value == cellcontcolor and ws.cell(i+1,4).fill.bgColor.value == cellcontcolor:
                    count_cont+=1
                    pair.update({commonname:{"skip":count, "content":count_cont}})
                    break
            df["CommonSetting"]="---"
            lists_2=list(pair.keys())      #common_setting名の配列
            k=0
            for i in range(len(lists_2)):
                #df.iat[k,len(df.columns)-1]=lists[i]
                skipnum = pair[lists_2[i]]["skip"]+pair[lists_2[i]]["content"]
                df.iloc[k:k+skipnum,len(df.columns)-1]=lists_2[i]
                k+=skipnum
            df=df.reindex(columns=[
                'CommonSetting','No','TestName', 'A-C', 'Process', 'Min', 'Typ', 'Max', 'Unit', 'Value',
                'Pin', 'Type', 'Parameter', 'Value.1', 'Unit.1', 'Pin.1', 'Measurement',
                'Parameter.1', 'Value.2', 'Unit.2', 'Type.1', 'Item', 'Ope', 'Add',
                'Data', 'MeasPoint', 'Comment', 'Comment.1'
            ])
            #-------------------------------------------------
            df=df.loc[df['TestName'].str.contains(search,case=False) & df['Process'].str.contains(process,case=False) & df['CommonSetting'].str.contains(searchcommon,case=False), :]
        context={"data":df.to_json(orient="records"),"name":sheet_read, 'cs':list(lists_2) }
                #'file_value':file_value,
        return JsonResponse(context,safe = False)


#-------------------------ここから下はグラフ描画のためのコードのため、現段階は不要。------------------------------------

def inputdata(pk):
    file_value=get_object_or_404(FileUpLoad,id=pk)
    try:
        df=pd.read_csv(file_value.upload_dir.path,skip_blank_lines=True,index_col=0,skiprows=6) 
    except UnicodeDecodeError:
        df=pd.read_csv(file_value.upload_dir.path,skip_blank_lines=True,index_col=0,encoding='cp932',skiprows=6)
    #df.drop(df.index[[1,2,3,4,5,6,7,8]])     #データフレームから余分な行を削除
    TestName="IREF"
    df.replace(' ','',regex=True,inplace=True)
    df.set_index('Test Item',inplace=True)
    Test = df.at[TestName,'data']
    n = len(Test)
    plt.hist(Test,bins=n,color="red")
    plt.xlabel("Current (uA)")
    plt.ylabel("Counts (-)")
    plt.xlim(Test.mean()-3*Test.std()*1.1,Test.mean()+3*Test.std()*1.1)
    plt.vlines(Test.mean()+3*Test.std(),0,max(Test),color="#5F9BFF",linestyle='dashed')
    plt.vlines(Test.mean()-3*Test.std(),0,max(Test),color="#5F9BFF",linestyle='dashed')

def conv():
    #ここのBytesIOで仮想メモリ確保している
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=200)
    s = buf.getvalue()
    #https://edge.sincar.jp/web/base64-inline-image/
    s=base64.b64encode(s).decode('utf-8')
    buf.close()
    return s

def histtests(req,pk):
    file_value=get_object_or_404(FileUpLoad,id=pk)
    inputdata(pk)            #plot図の作成
    response = conv()        #png形式に変換する。
    #response = HttpResponse(png, content_type='image/png')
    context={
            'file_value':file_value,
            'response':response,
    }
    plt.cla()        #いちおう初期化する。
    return render(req,'histtest.html',context)
    